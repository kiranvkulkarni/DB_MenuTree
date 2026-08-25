"""Build a workbook shaped like the hand-authored MenuTree specification.

Mirrors the real deliverable's layout so the reader can be tested without
the actual sheet: a summary block at the top, a header row carrying
`Test Result / Defect ID / Comments` followed by `1 Depth ... N Depth`, and
one row per element with its label in the column matching its depth.

Content is modelled on the Modes and Settings sheets, including the awkward
cases: [bracketed] context rows, `[Title]` / `(On/Off)` annotations, and a
permission dialog whose branches sit under it.

    python tests/make_spec_fixture.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

OUT = Path("tests/spec_fixture.xlsx")

# (depth, label) in the sheet's own top-to-bottom tree order.
MODES = [
    (1, "Camera"),
    (2, "[First entry camera]"),
    (2, "Location tags [ON/OFF] on 3rd entry of camera"),
    (3, "[When location Permission is OFF in Dut]"),
    (3, "Turn on location tags[Title]"),
    (3, "Learn more"),
    (4, "Location tags [Title]"),
    (4, "Location tags ...recorded."),
    (4, "OK"),
    (3, "Cancel"),
    (3, "Turn on"),
    (4, "To continue, Turn on device location"),
    (4, "No Thanks"),
    (4, "OK"),
    (3, "Precise"),
    (3, "Approximate"),
    (4, "while using this app"),
    (4, "Only this time"),
    (4, "Don't Allow"),
    (2, "[Rear Camera]"),
    (2, "Photo"),
    (3, "Photo [Title]"),
    (3, "Flash icon"),
    (4, "On"),
    (4, "Off"),
    (4, "Auto"),
    (3, "Resolution"),
    (4, "12MP"),
    (4, "24MP"),
    (4, "50MP"),
    (4, "200MP"),
    (3, "Motion Photos"),
    (4, "ON"),
    (4, "OFF"),
    (3, "Filters"),
    (4, "Original"),
    (4, "Classic film"),
    (4, "Crystal"),
    (4, "Blanc"),
]

SETTINGS = [
    (1, "Camera"),
    (2, "Quick Control"),
    (3, "Settings icon"),
    (4, "Back Icon"),
    (4, "Camera Settings [Title]"),
    (4, "Intelligent features [subtitle]"),
    (4, "Scanning"),
    (5, "Scan Document and text (On/Off)"),
    (5, "Scan Text(on/off)"),
    (5, "Scan QR codes(on/off)"),
    (5, "Back key icon"),
    (4, "Shot Suggestion"),
    (5, "ON"),
    (5, "OFF"),
    (4, "Photo enhance"),
    (5, "Priorize quality (Radio button On/Off)"),
    (5, "Balance (Radio button On/Off)"),
    (5, "Priorize speed (Radio button On/Off)"),
    (5, "Scene Optimiser (On/Off)"),
    (5, "Back key icon"),
    (4, "Photos (sub-title)"),
    (4, "Photo Format"),
    (5, "High efficiency pitures(ON/OFF)"),
    (5, "Pro mode piture format"),
    (6, "JEPG format"),
    (6, "RAW and JEPG format"),
    (6, "RAW format"),
    (6, "Back key  icon"),
    (4, "Watermark(On/Off)"),
    (5, "ON"),
    (5, "select information that include watermark"),
    (6, "Model name"),
    (7, "edit"),
    (6, "Date"),
    (6, "Time"),
]


def build_sheet(book, title: str, entries, max_depth: int, first: bool):
    sheet = book.active if first else book.create_sheet()
    sheet.title = title

    # Summary block, as in the real workbook.
    sheet.cell(row=2, column=2, value="PRELOAD")
    sheet.cell(row=2, column=5, value="Total")
    sheet.cell(row=2, column=7, value="Pass")
    sheet.cell(row=2, column=9, value="Fail")
    sheet.cell(row=2, column=11, value="NA")
    sheet.cell(row=2, column=13, value="NT")
    sheet.cell(row=3, column=2, value="PASS")
    sheet.cell(row=3, column=5, value=len(entries))

    header_row = 5
    sheet.cell(row=header_row, column=1, value="Test Result")
    sheet.cell(row=header_row, column=2, value="Defect ID")
    sheet.cell(row=header_row, column=3, value="Comments")
    for depth in range(1, max_depth + 1):
        sheet.cell(row=header_row, column=3 + depth, value=f"{depth} Depth")

    for offset, (depth, label) in enumerate(entries):
        row = header_row + 1 + offset
        sheet.cell(row=row, column=1, value="Pass")
        sheet.cell(row=row, column=3 + depth, value=label)
    return sheet


def main() -> int:
    from openpyxl import Workbook

    book = Workbook()
    build_sheet(book, "Modes", MODES, 18, first=True)
    build_sheet(book, "Settings", SETTINGS, 7, first=False)

    # A sheet with no depth columns, to prove it is skipped rather than
    # misread -- the real workbook has a Summary tab like this.
    summary = book.create_sheet()
    summary.title = "Summary"
    summary.cell(row=2, column=2, value="PROJECT INFORMATION")
    summary.cell(row=4, column=2, value="Test result")
    summary.cell(row=5, column=2, value="Pass")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    book.save(OUT)
    print(f"wrote {OUT} ({len(MODES)} Modes + {len(SETTINGS)} Settings rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
