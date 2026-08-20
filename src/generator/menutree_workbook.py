"""Export the MenuTree as an Excel workbook.

One sheet, laid out like the existing deliverable: a small result summary at
the top, then one row per element with its label in the column matching its
depth. The final column carries the UVTA test case for that row, so the
navigation path, the assertion and the result all sit on the same line.

Test Result defaults to `NT` (not tested). This tool discovers and records
the tree; it does not execute it. Marking rows `Pass` would assert something
never checked.
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional, Sequence

logger = logging.getLogger(__name__)

RESULT_NOT_TESTED = "NT"
RESULTS = ("Pass", "Fail", "NA", "NT")

_HEADER_FILL = "1F4E79"
_SUMMARY_FILL = "DDEBF7"
_RESULT_FILLS = {
    "Pass": "C6EFCE",
    "Fail": "FFC7CE",
    "NA": "FFEB9C",
    "NT": "F2F2F2",
}


def _autosize(sheet, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            first_line = str(cell.value).split("\n")[0]
            widths[cell.column] = max(widths.get(cell.column, 0), len(first_line))
    from openpyxl.utils import get_column_letter
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(10, width + 2), max_width
        )


def write_workbook(
    rows: Sequence[Dict],
    path: Path,
    package: str,
    max_depth: int = 0,
    uvta_by_row: Optional[Dict[int, str]] = None,
    result_default: str = RESULT_NOT_TESTED,
    stats: Optional[Dict] = None,
) -> Path:
    """Write a single-sheet MenuTree workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    if not rows:
        raise ValueError("refusing to write an empty MenuTree workbook")

    depth = max_depth or max(int(r.get("depth", 1)) for r in rows)
    uvta_by_row = uvta_by_row or {}

    book = Workbook()
    sheet = book.active
    sheet.title = "MenuTree"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    top = Alignment(vertical="top", wrap_text=True)

    # -- summary ---------------------------------------------------------
    counts = {r: 0 for r in RESULTS}
    counts[result_default] = len(rows)
    summary = [
        ("Package", package),
        ("Total", len(rows)),
        ("Pass", counts["Pass"]),
        ("Fail", counts["Fail"]),
        ("NA", counts["NA"]),
        ("NT", counts["NT"]),
        ("Max depth", depth),
    ]
    for index, (label, value) in enumerate(summary, start=1):
        key = sheet.cell(row=1, column=index, value=label)
        key.font = Font(bold=True)
        key.fill = PatternFill("solid", fgColor=_SUMMARY_FILL)
        sheet.cell(row=2, column=index, value=value)

    if stats:
        note = (
            f"discovered in {stats.get('elapsed_seconds', '?')}s; "
            f"screens={stats.get('screens_visited', '?')}, "
            f"descents={stats.get('descents', '?')}, "
            f"guard blocked={((stats.get('guard') or {}).get('blocked_attempts', 0))}"
        )
        sheet.cell(row=3, column=1, value=note).font = Font(italic=True, size=9)

    # -- header ----------------------------------------------------------
    header_row = 5
    header = ["Test Result", "Defect ID", "Comments"]
    header += [f"{i} Depth" for i in range(1, depth + 1)]
    header += ["Kind", "UVTA Test Case"]

    for index, title in enumerate(header, start=1):
        cell = sheet.cell(row=header_row, column=index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # -- data ------------------------------------------------------------
    uvta_column = len(header)
    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset

        result = sheet.cell(row=excel_row, column=1, value=result_default)
        result.fill = PatternFill("solid", fgColor=_RESULT_FILLS[result_default])
        result.alignment = Alignment(horizontal="center")

        sheet.cell(row=excel_row, column=2, value="")

        comment = row.get("note", "") or ""
        if row.get("blocked"):
            comment = (
                f"blocked by action guard ({row['blocked']}); route not walked"
            )
        sheet.cell(row=excel_row, column=3, value=comment).alignment = top

        column = 3 + max(1, min(int(row.get("depth", 1)), depth))
        label_cell = sheet.cell(row=excel_row, column=column, value=row.get("label", ""))
        label_cell.alignment = top

        sheet.cell(row=excel_row, column=3 + depth + 1, value=row.get("kind", ""))

        uvta = uvta_by_row.get(offset, "")
        uvta_cell = sheet.cell(row=excel_row, column=uvta_column, value=uvta)
        uvta_cell.alignment = top
        uvta_cell.font = Font(name="Consolas", size=9)

    # A dropdown makes the sheet usable as a test record straight away.
    validation = DataValidation(
        type="list", formula1=f'"{",".join(RESULTS)}"', allow_blank=True
    )
    sheet.add_data_validation(validation)
    validation.add(f"A{header_row + 1}:A{header_row + len(rows)}")

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=4)
    sheet.auto_filter.ref = (
        f"A{header_row}:"
        f"{sheet.cell(row=header_row, column=uvta_column).coordinate}"
    )
    _autosize(sheet)
    sheet.column_dimensions[
        sheet.cell(row=header_row, column=uvta_column).column_letter
    ].width = 60

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    logger.info(
        "MenuTree workbook written: %s (%d rows, %d depth columns)",
        path.resolve(), len(rows), depth,
    )
    return path
