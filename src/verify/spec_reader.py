"""Read a hand-authored MenuTree workbook as the expected specification.

The sheet is the source of truth: 1,896 rows, one element per row, its label
placed in the column matching its depth, alongside Test Result / Defect ID /
Comments. This module turns that layout back into a tree.

Two properties of the layout make this reliable:

* **Depth is positional.** A row's depth is which `N Depth` column holds its
  label, so the tree structure is explicit rather than inferred.
* **Rows are in tree order.** The sheet is a depth-first listing, so a row's
  parent is the nearest preceding row one level shallower. No ids needed.

Nothing here assumes a particular sheet name or a fixed header row -- the
header is located by looking for the depth columns, so minor layout drift
between workbooks does not break it.
"""
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence

logger = logging.getLogger(__name__)

DEPTH_HEADER = re.compile(r"^\s*(\d+)\s*depth\s*$", re.IGNORECASE)

# Human annotations appended for readers. They are not on-screen text, so they
# must be stripped before a label is used as a selector.
ANNOTATION = re.compile(
    r"\s*(?:"
    r"\[\s*title\s*\]"
    r"|\[\s*sub-?title\s*\]|\(\s*sub-?title\s*\)"
    r"|\[\s*on\s*/\s*off\s*\]|\(\s*on\s*/\s*off\s*\)"
    r"|\(\s*radio button\s+on\s*/\s*off\s*\)"
    r")\s*",
    re.IGNORECASE,
)

# A bracketed row that names a context or precondition rather than a control:
#   [First entry camera]   [Rear Camera]   [When location Permission is OFF]
# These cannot be clicked or asserted, so they are carried as context for the
# rows beneath them rather than verified.
CONTEXT_ROW = re.compile(r"^\s*\[.+\]\s*$")

RESULT_COLUMNS = {
    "test result": "result",
    "result": "result",
    "defect id": "defect_id",
    "defect": "defect_id",
    "comments": "comments",
    "comment": "comments",
}


@dataclass
class SpecRow:
    """One expected element, as written in the workbook."""
    sheet: str
    excel_row: int          # 1-based, so results can be written back in place
    depth: int
    label: str              # exactly as authored, annotations included
    selector_text: str      # annotation-stripped, usable against the screen
    path: List[str] = field(default_factory=list)      # ancestor selector texts
    context: List[str] = field(default_factory=list)   # enclosing [bracketed] rows
    is_context: bool = False
    is_root: bool = False
    authored_result: str = ""
    comments: str = ""

    @property
    def key(self) -> str:
        return f"{self.sheet}!{self.excel_row}"


def strip_annotations(label: str) -> str:
    """The on-screen text a selector should match."""
    return ANNOTATION.sub(" ", label or "").strip()


def _find_header(sheet, scan_rows: int = 30) -> Optional[Dict]:
    """Locate the header row and map its columns.

    Found by looking for the `N Depth` columns rather than a fixed row index,
    so a workbook with a taller or shorter summary block still reads.
    """
    for row_index in range(1, min(scan_rows, sheet.max_row) + 1):
        depth_columns: Dict[int, int] = {}
        meta_columns: Dict[str, int] = {}
        for cell in sheet[row_index]:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            match = DEPTH_HEADER.match(text)
            if match:
                depth_columns[int(match.group(1))] = cell.column
                continue
            key = RESULT_COLUMNS.get(text.lower())
            if key:
                meta_columns[key] = cell.column
        if len(depth_columns) >= 2:
            return {
                "row": row_index,
                "depth_columns": dict(sorted(depth_columns.items())),
                "meta_columns": meta_columns,
            }
    return None


def read_sheet(sheet) -> List[SpecRow]:
    header = _find_header(sheet)
    if not header:
        logger.info("sheet %r has no depth columns; skipped", sheet.title)
        return []

    depth_columns = header["depth_columns"]
    meta = header["meta_columns"]
    rows: List[SpecRow] = []

    # Nearest ancestor at each depth, so a row's path is read off directly.
    ancestors: Dict[int, str] = {}
    context_stack: Dict[int, str] = {}

    for excel_row in range(header["row"] + 1, sheet.max_row + 1):
        depth = None
        label = ""
        for level, column in depth_columns.items():
            value = sheet.cell(row=excel_row, column=column).value
            if value is not None and str(value).strip():
                depth, label = level, str(value).strip()
                break
        if depth is None:
            continue

        is_context = bool(CONTEXT_ROW.match(label))
        # Depth 1 names the application, not a control on a screen. Launching
        # the app satisfies it, so it must never become a click step -- left
        # in, every path began with a "Camera" tap that matches nothing and
        # every row failed with "path step not found on screen".
        is_root = depth == 1
        selector_text = strip_annotations(label)

        # Anything at or below this row's depth stops being an ancestor.
        for level in [d for d in ancestors if d >= depth]:
            ancestors.pop(level, None)

        # Context is scoped differently: a bracketed marker applies to its
        # *siblings* as well as its descendants. "[When location Permission
        # is OFF in Dut]" sits at the same depth as the Cancel / Turn on rows
        # it qualifies, so dropping same-depth context would detach the
        # precondition from the rows that need it. Only a deeper marker is
        # discarded here; a same-depth one is replaced below when another
        # bracketed row appears at that level.
        for level in [d for d in context_stack if d > depth]:
            context_stack.pop(level, None)
        if is_context:
            context_stack.pop(depth, None)

        path = [ancestors[d] for d in sorted(ancestors) if d < depth]
        context = [context_stack[d] for d in sorted(context_stack) if d <= depth]

        rows.append(SpecRow(
            sheet=sheet.title,
            excel_row=excel_row,
            depth=depth,
            label=label,
            selector_text=selector_text,
            path=path,
            context=context,
            is_context=is_context,
            is_root=is_root,
            authored_result=str(
                sheet.cell(row=excel_row, column=meta["result"]).value or ""
            ).strip() if "result" in meta else "",
            comments=str(
                sheet.cell(row=excel_row, column=meta["comments"]).value or ""
            ).strip() if "comments" in meta else "",
        ))

        if is_context:
            context_stack[depth] = label
        elif not is_root:
            ancestors[depth] = selector_text

    logger.info("sheet %r: %d spec row(s), max depth %d",
                sheet.title, len(rows),
                max((r.depth for r in rows), default=0))
    return rows


def read_workbook(path: Path, sheets: Optional[Sequence[str]] = None) -> List[SpecRow]:
    """Read every sheet that looks like a MenuTree listing."""
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=False)
    out: List[SpecRow] = []
    for name in book.sheetnames:
        if sheets and name not in sheets:
            continue
        out.extend(read_sheet(book[name]))
    if not out:
        raise ValueError(
            f"No MenuTree rows found in {path}. Expected columns named "
            "'1 Depth', '2 Depth', ... on some sheet."
        )
    logger.info("%s: %d spec row(s) across %d sheet(s)",
                path.name, len(out), len({r.sheet for r in out}))
    return out


def inspect_workbook(path: Path,
                     sheets: Optional[Sequence[str]] = None) -> List[Dict]:
    """Per-sheet parse diagnostics: what was found, and what was skipped.

    Exists to answer one question without seeing the workbook: **are rows
    being dropped?** A row is only read if one of the `N Depth` columns holds
    its label, so if the header scan finds seven depth columns on a sheet
    that actually has eighteen, every row living in columns 8-18 vanishes
    silently -- and the result looks like a smaller, shallower app rather
    than a parse failure.

    Reports no cell contents, only counts and column positions.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=False)
    out: List[Dict] = []
    for name in book.sheetnames:
        if sheets and name not in sheets:
            continue
        sheet = book[name]
        header = _find_header(sheet)
        if not header:
            out.append({"sheet": name, "header_row": None, "depth_columns": [],
                        "read": 0, "skipped": 0, "last_row": sheet.max_row})
            continue

        depth_columns = header["depth_columns"]
        read = skipped = 0
        first_skipped: List[int] = []
        for excel_row in range(header["row"] + 1, sheet.max_row + 1):
            has_depth = any(
                (sheet.cell(row=excel_row, column=col).value is not None
                 and str(sheet.cell(row=excel_row, column=col).value).strip())
                for col in depth_columns.values()
            )
            if has_depth:
                read += 1
                continue
            # A row with content but nothing in any depth column: either a
            # spacer, or a row whose label sits in a column we never found.
            if any(c.value is not None and str(c.value).strip()
                   for c in sheet[excel_row]):
                skipped += 1
                if len(first_skipped) < 10:
                    first_skipped.append(excel_row)

        out.append({
            "sheet": name,
            "header_row": header["row"],
            "depth_columns": sorted(depth_columns),
            "depth_column_letters": [
                sheet.cell(row=header["row"], column=c).column_letter
                for _, c in sorted(depth_columns.items())
            ],
            "read": read,
            "skipped": skipped,
            "first_skipped_rows": first_skipped,
            "last_row": sheet.max_row,
        })
    return out


def diagnostics(infos: Sequence[Dict]) -> str:
    """Render inspect_workbook output. Counts and columns only."""
    lines = ["", "=" * 60, "  PARSE DIAGNOSTICS (numbers only -- safe to share)",
             "=" * 60]
    for info in infos:
        lines.append(f"  sheet {info['sheet']!r}")
        if info["header_row"] is None:
            lines.append("      no depth columns found -- sheet skipped entirely")
            continue
        depths = info["depth_columns"]
        lines.append(f"      header row          : {info['header_row']}")
        lines.append(f"      depth columns found : {len(depths)}  "
                     f"({'N Depth for N=' + ','.join(map(str, depths)) if depths else 'none'})")
        lines.append(f"      spreadsheet columns : "
                     f"{','.join(info.get('depth_column_letters') or [])}")
        lines.append(f"      last row in sheet   : {info['last_row']}")
        lines.append(f"      rows READ           : {info['read']}")
        lines.append(f"      rows SKIPPED        : {info['skipped']}"
                     "   <- non-empty rows with nothing in any depth column")
        if info["skipped"]:
            lines.append("          first few at rows: "
                         + ", ".join(map(str, info["first_skipped_rows"])))
            lines.append("          If these are real menu entries, a depth")
            lines.append("          column was missed -- tell me which columns")
            lines.append("          hold them and how the header names them.")
    lines.append("=" * 60)
    return "\n".join(lines)


def health(rows: Sequence[SpecRow]) -> str:
    """Structural check of a parsed workbook, in numbers only.

    Deliberately prints **no labels**. The real workbook cannot leave the
    infrastructure it lives on, so this is what can be read aloud, pasted
    into a chat, or attached to a ticket without disclosing its content --
    counts, shapes, and the row numbers of anything that looks wrong, so a
    person with the sheet open can go and look.

    Everything the verifier does rests on the reconstructed `path` being
    right, and a path is only as good as the layout assumptions in this
    module. These are the ways those assumptions fail.
    """
    real = [r for r in rows if not r.is_context]
    lines = ["", "=" * 60, "  SPEC HEALTH (numbers only -- safe to share)", "=" * 60,
             f"  rows parsed           : {len(rows)}",
             f"  context rows          : {len(rows) - len(real)}",
             f"  max depth             : {max((r.depth for r in rows), default=0)}"]

    # A row deeper than 1 with no reconstructed path has no parent: either
    # the header was misread or the sheet is not in depth-first order.
    # Depth 1 is the app and depth 2 sits directly under it, so an empty
    # path is correct for both. Only depth 3+ needs ancestors.
    orphans = [r for r in real if r.depth > 2 and not r.path]
    lines.append(f"  rows with NO path     : {len(orphans)}   <- must be 0")
    if orphans:
        lines.append("      first few at rows: "
                     + ", ".join(f"{r.sheet}!{r.excel_row}" for r in orphans[:8]))

    # Path length should be depth-1 for a well-formed tree (root excluded).
    # Depth 1 = the app (no path). Depth 2 = one click from launch, path [].
    # So a well-formed row at depth d carries d-2 ancestors.
    wrong = [r for r in real
             if not r.is_root and len(r.path) != r.depth - 2]
    lines.append(f"  path length mismatch  : {len(wrong)}   <- expect 0")
    if wrong:
        lines.append("      first few at rows: "
                     + ", ".join(f"{r.sheet}!{r.excel_row}(d{r.depth},p{len(r.path)})"
                                 for r in wrong[:8]))

    # An empty selector means the whole label was annotation.
    empty = [r for r in real if not r.selector_text]
    lines.append(f"  empty selector text   : {len(empty)}   <- these can never match")
    if empty:
        lines.append("      first few at rows: "
                     + ", ".join(f"{r.sheet}!{r.excel_row}" for r in empty[:8]))

    # Two siblings with the same selector are indistinguishable on screen.
    seen: Dict[tuple, int] = {}
    for r in real:
        key = (r.sheet, tuple(r.path), r.selector_text.lower())
        seen[key] = seen.get(key, 0) + 1
    dupes = sum(n - 1 for n in seen.values() if n > 1)
    lines.append(f"  duplicate siblings    : {dupes}   <- ambiguous, first match wins")

    lines.append(f"  rows needing a human  : "
                 f"{sum(1 for r in real if r.context)}   (have a precondition)")

    lines.append("-" * 60)
    lines.append("  path depth distribution (how far the walk must descend):")
    by_len: Dict[int, int] = {}
    for r in real:
        by_len[len(r.path)] = by_len.get(len(r.path), 0) + 1
    for n in sorted(by_len):
        lines.append(f"    {n} step(s) from launch : {by_len[n]}")
    lines.append("=" * 60)
    return "\n".join(lines)


def summarise(rows: Sequence[SpecRow]) -> str:
    by_sheet: Dict[str, int] = {}
    by_depth: Dict[int, int] = {}
    for row in rows:
        by_sheet[row.sheet] = by_sheet.get(row.sheet, 0) + 1
        by_depth[row.depth] = by_depth.get(row.depth, 0) + 1
    context = sum(1 for r in rows if r.is_context)

    lines = ["", "=" * 58, "  EXPECTED SPECIFICATION", "=" * 58,
             f"  rows            : {len(rows)}",
             f"  context rows    : {context} (preconditions, not verified)",
             f"  verifiable rows : {len(rows) - context}",
             f"  max depth       : {max(by_depth) if by_depth else 0}",
             "-" * 58, "  by sheet:"]
    for name in sorted(by_sheet):
        lines.append(f"    {name:<24} {by_sheet[name]}")
    lines.append("-" * 58)
    lines.append("  by depth:")
    for depth in sorted(by_depth):
        lines.append(f"    depth {depth:<3} {by_depth[depth]}")
    lines.append("=" * 58)
    return "\n".join(lines)
