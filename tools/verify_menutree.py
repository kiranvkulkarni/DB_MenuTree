"""Verify a build against the hand-authored MenuTree workbook.

    python tools/verify_menutree.py --spec MenuTree.xlsx \
        --package com.oplus.camera --serial <serial>

Reads the expected tree from the workbook, walks it on the device, and
writes a *copy* with Test Result / Comments filled in. The original is never
modified.

    --dry-run   parse the spec and report what would be checked, no device
"""
import argparse
import json
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.crawler.action_guard import DEFAULT_PRESETS, GUARD_PRESETS  # noqa: E402
from src.logging_setup import setup_logging  # noqa: E402
from src.run_lock import DeviceBusy, RunLock  # noqa: E402
from src.verify.spec_reader import (  # noqa: E402
    diagnostics,
    health,
    inspect_workbook,
    read_workbook,
    summarise,
)
from src.verify.verifier import FAIL, NA, PASS, MenuTreeVerifier  # noqa: E402

logger = logging.getLogger("verify_menutree")

_FILLS = {PASS: "C6EFCE", FAIL: "FFC7CE", NA: "FFEB9C", "NT": "F2F2F2"}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Verify a build against the MenuTree spec")
    p.add_argument("--spec", required=True, help="the hand-authored workbook")
    p.add_argument("--package", required=True)
    p.add_argument("--serial", default=None)
    p.add_argument("--sheets", default="",
                   help="comma-separated sheet names; default is every sheet "
                        "that has depth columns")
    p.add_argument("--output-root", default="./output")
    p.add_argument("--time-budget", type=float, default=7200)
    p.add_argument("--ready-timeout", type=float, default=20.0)
    p.add_argument("--settle", type=float, default=1.0)
    p.add_argument("--driver", default="auto", choices=("auto", "u2", "adb"))
    p.add_argument("--no-reset", action="store_true",
                   help="do not pm clear before verifying")
    p.add_argument("--no-guard", action="store_true")
    p.add_argument("--guard-presets", default=",".join(DEFAULT_PRESETS),
                   help=f"available: {', '.join(sorted(GUARD_PRESETS))}")
    p.add_argument("--guard-extra", default="")
    p.add_argument("--dry-run", action="store_true",
                   help="parse and report the spec only; no device needed")
    p.add_argument("--force-lock", action="store_true",
                   help="start even if another run holds this device")
    return p.parse_args()


def write_results(spec_path: Path, out_path: Path, results) -> Path:
    """Copy the workbook and fill in Test Result / Comments per row."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    shutil.copyfile(spec_path, out_path)
    book = load_workbook(out_path)

    # Locate each sheet's Test Result / Comments columns once.
    columns = {}
    for name in book.sheetnames:
        sheet = book[name]
        found = {}
        for row_index in range(1, min(30, sheet.max_row) + 1):
            for cell in sheet[row_index]:
                text = str(cell.value).strip().lower() if cell.value else ""
                if text in ("test result", "result"):
                    found["result"] = cell.column
                elif text in ("comments", "comment"):
                    found["comments"] = cell.column
            if "result" in found:
                break
        columns[name] = found

    for item in results:
        sheet = book[item.row.sheet]
        found = columns.get(item.row.sheet, {})
        if "result" in found:
            cell = sheet.cell(row=item.row.excel_row, column=found["result"],
                              value=item.result)
            cell.fill = PatternFill("solid", fgColor=_FILLS.get(item.result, "F2F2F2"))
            if item.result == FAIL:
                cell.font = Font(bold=True)
        if "comments" in found and item.detail:
            sheet.cell(row=item.row.excel_row, column=found["comments"],
                       value=item.detail)

    book.save(out_path)
    return out_path


def main() -> int:
    args = parse_args()
    spec_path = Path(args.spec)
    if not spec_path.exists():
        print(f"spec not found: {spec_path}")
        return 1

    now = datetime.now()
    stamp = f"{now:%Y%m%d_%H%M%S}_{now.microsecond // 1000:03d}"
    output_dir = Path(args.output_root) / f"verify_{args.package}_{stamp}"
    output_dir.mkdir(parents=True, exist_ok=True)
    setup_logging(log_dir=str(output_dir), level="INFO", run_id=output_dir.name)
    logger.info("Run folder: %s", output_dir.resolve())

    sheets = [s.strip() for s in args.sheets.split(",") if s.strip()] or None
    spec = read_workbook(spec_path, sheets)
    print(summarise(spec))

    if args.dry_run:
        print(health(spec))
        print(diagnostics(inspect_workbook(spec_path, sheets)))
        preview = output_dir / "spec_parsed.json"
        preview.write_text(json.dumps(
            [{"sheet": r.sheet, "excel_row": r.excel_row, "depth": r.depth,
              "label": r.label, "selector_text": r.selector_text,
              "path": r.path, "context": r.context,
              "is_context": r.is_context} for r in spec],
            indent=2), encoding="utf-8")
        print(f"\n  dry run -- parsed spec written to {preview.resolve()}")
        print()
        print("  NOTE: spec_parsed.json contains your sheet's labels. The")
        print("  SPEC HEALTH block above is numbers only and safe to share;")
        print("  the JSON is not.")
        print("  Check the paths and selector_text against the real UI, then "
              "re-run without --dry-run.")
        return 0

    try:
        lock = RunLock(args.output_root, args.serial).acquire(args.force_lock)
    except DeviceBusy as exc:
        logger.error("%s", exc)
        print()
        print(f"  {exc}")
        print()
        return 1

    verifier = MenuTreeVerifier(args.package, args.serial, {
        "time_budget": args.time_budget,
        "ready_timeout": args.ready_timeout,
        "settle_seconds": args.settle,
        "driver": args.driver,
        "reset_before_start": not args.no_reset,
        "guard_enabled": not args.no_guard,
        "guard_presets": [g for g in args.guard_presets.split(",") if g],
        "guard_extra_patterns": [g for g in args.guard_extra.split(",") if g],
    })

    try:
        report = verifier.verify(spec)
    except KeyboardInterrupt:
        logger.warning("Interrupted; writing what was verified so far.")
        report = getattr(verifier, "_partial", None)
        if report is None:
            return 1
    finally:
        verifier._release()
        lock.release()

    (output_dir / "verify_results.json").write_text(
        json.dumps({"package": args.package,
                    "stats": report.stats,
                    "results": [r.to_dict() for r in report.results]},
                   indent=2), encoding="utf-8")

    filled = write_results(
        spec_path, output_dir / f"{spec_path.stem}_verified.xlsx", report.results)

    counts = report.counts()
    print()
    print("=" * 58)
    print("  VERIFICATION RESULT")
    print("=" * 58)
    print(f"  spec rows      : {report.stats['rows_in_spec']}")
    print(f"  checked        : {report.stats['rows_checked']}")
    for name in (PASS, FAIL, NA, "NT"):
        if counts.get(name):
            print(f"  {name:<15}: {counts[name]}")
    print("-" * 58)
    print(f"  PASS RATE      : {report.pass_rate()}%  (of Pass+Fail)")
    print(f"  navigations    : {report.stats['navigations']}")
    print(f"  relaunches     : {report.stats['relaunches']}")
    print(f"  elapsed        : {report.stats['elapsed_seconds']}s")
    print("=" * 58)

    failures = [r for r in report.results if r.result == FAIL]
    if failures:
        print(f"\n  {len(failures)} FAILING ROW(S):")
        for item in failures[:25]:
            where = " > ".join(item.row.path[-2:]) or "<root>"
            print(f"    {item.row.sheet}!{item.row.excel_row}  "
                  f"{where} > {item.row.label[:38]}")
            print(f"        {item.detail}")
        if len(failures) > 25:
            print(f"    ... and {len(failures) - 25} more")

    print(f"\n  workbook: {filled.resolve()}")
    return 2 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
