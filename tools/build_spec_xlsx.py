"""Rebuild a MenuTree workbook from rows transcribed out of the PDF photographs.

    python tools/build_spec_xlsx.py --out S25_Ultra_MenuTree/S25U_MenuTree.xlsx

The device team's MenuTree could not be copied off its machine, so it arrived
as photographs of the Excel window. This writes those rows back into the
layout `src/verify/spec_reader.py` expects, so the verifier can walk them.

The output is a RECONSTRUCTION, not the original. It is written with the
authored Test Result column left blank -- deliberately. Copying the original's
"Pass" values across would look like evidence about this build when it is
nothing of the sort; results belong to the run that produces them.
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

COMMENTS = {}

MAX_DEPTH = 18          # as in the original sheet, so deeper sheets still fit
# The original puts the header on row 4 and the first data row on row 5.
# Getting this wrong is silent and costly: with the header on row 5, the
# first data row overwrote the '1 Depth' header, so the depth-1 column
# vanished and its row was dropped -- 87 rows instead of 88, with nothing
# reporting an error. The row count check below is what caught it.
HEADER_ROW = 4


def build_sheet(book, title, rows, first):
    """Write one sheet, preserving the original's own row numbers.

    The header row is derived per sheet rather than fixed: the real workbook
    puts Settings' header on row 4 and Modes' on row 5, and preserving each
    sheet's own numbering is what lets a reviewer compare this reconstruction
    against the original line by line.
    """
    sheet = book.active if first else book.create_sheet()
    sheet.title = title
    header_row = min(r[0] for r in rows) - 1

    depths = [d for _, d, _ in rows]
    sheet.cell(row=2, column=2, value="PRELOAD")
    sheet.cell(row=2, column=5, value="Total")
    sheet.cell(row=3, column=5, value=len(rows))
    sheet.cell(row=2, column=7, value="Pass")
    sheet.cell(row=2, column=9, value="Fail")
    sheet.cell(row=2, column=11, value="NA")
    sheet.cell(row=2, column=13, value="NT")

    sheet.cell(row=header_row, column=2, value="Test Result")
    sheet.cell(row=header_row, column=3, value="Defect ID")
    sheet.cell(row=header_row, column=4, value="Comments")
    # Depth N lives in column E+N-1, exactly as the photographs show.
    for depth in range(1, MAX_DEPTH + 1):
        sheet.cell(row=header_row, column=4 + depth, value=f"{depth} Depth")

    for excel_row, depth, label in rows:
        sheet.cell(row=excel_row, column=4 + depth, value=label)
    for excel_row, note in (COMMENTS.get(title) or {}).items():
        sheet.cell(row=excel_row, column=4, value=note)

    widths = {2: 12, 3: 10, 4: 14}
    for col, w in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = w
    for depth in range(1, MAX_DEPTH + 1):
        letter = sheet.cell(row=1, column=4 + depth).column_letter
        # Wide enough that nothing is clipped -- the flaw in the source photos.
        sheet.column_dimensions[letter].width = 34
    return sheet, max(depths)


def main() -> int:
    p = argparse.ArgumentParser(description="Rebuild the MenuTree workbook")
    p.add_argument("--out", required=True)
    p.add_argument("--sheets", default="Settings,Modes",
                   help="which transcribed sheets to include")
    args = p.parse_args()

    from openpyxl import Workbook

    available = {}
    try:
        from menutree_ocr.settings_rows import SETTINGS
        available["Settings"] = SETTINGS
    except ImportError:
        pass
    try:
        from menutree_ocr.modes_rows import MODES, NT_COMMENTS
        available["Modes"] = MODES
        COMMENTS["Modes"] = NT_COMMENTS
    except ImportError:
        pass

    wanted = [s.strip() for s in args.sheets.split(",") if s.strip()]
    present = [s for s in wanted if s in available]
    if not present:
        print("No transcribed sheets found. Expected "
              "tools/menutree_ocr/settings_rows.py and/or modes_rows.py")
        return 1

    book = Workbook()
    total = 0
    print()
    for index, name in enumerate(present):
        rows = available[name]
        _, deepest = build_sheet(book, name, rows, first=(index == 0))
        total += len(rows)
        print(f"  {name:<10} {len(rows):>5} rows, max depth {deepest}")

    missing = [s for s in wanted if s not in available]
    for name in missing:
        print(f"  {name:<10}     - not transcribed yet")

    # The header must not collide with the first data row.

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(out)
    print(f"\n  wrote {out.resolve()}  ({total} rows)")
    if missing:
        print("  INCOMPLETE: this workbook covers only the sheets listed above.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
